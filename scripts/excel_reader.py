from openpyxl import load_workbook

REQUIRED_COLS = [
    "TeamID",
    "Team Name",
    "GitHub Usernames"
]

def read_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]

    for col in REQUIRED_COLS:
        if col not in headers:
            raise ValueError(f"Missing column: {col}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        if not data["TeamID"] or not data["Team Name"]:
            raise ValueError("TeamID and Team Name are mandatory")

        if not data["GitHub Usernames"]:
            raise ValueError(f"No GitHub usernames for {data['TeamID']}")

        rows.append(data)

    return rows
